"""
Reports router — Excel/PDF export, financial summaries.
"""
from datetime import date
from io import BytesIO
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from app.core.database import get_db
from app.core.deps import require_roles
from app.models.user import User, UserRole
from app.models.member import Member
from app.models.payment import Payment, PaymentStatus
from app.models.expense import Expense
from app.models.subscription import Subscription, SubscriptionStatus

router = APIRouter(prefix="/reports", tags=["Reports"])
REPORT_ROLES = [UserRole.SUPER_ADMIN, UserRole.MANAGER, UserRole.ACCOUNTANT]


@router.get("/members/export")
async def export_members_excel(db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(REPORT_ROLES))):
    """Export all members to Excel."""
    from openpyxl import Workbook
    result = await db.execute(select(Member).order_by(Member.created_at.desc()))
    members = result.scalars().all()
    wb = Workbook(); ws = wb.active; ws.title = "Members"
    ws.append(["Name (AR)", "Name (EN)", "Phone", "National ID", "Gender", "Status", "Created"])
    for m in members:
        ws.append([m.name_ar, m.name_en or "", m.phone, m.national_id or "",
            m.gender.value, "Active" if m.is_active else "Inactive", str(m.created_at.date())])
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=members.xlsx"})


@router.get("/financial/export")
async def export_financial_excel(
    month: int = Query(..., ge=1, le=12), year: int = Query(..., ge=2020),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(REPORT_ROLES))):
    """Export monthly financial report to Excel."""
    from openpyxl import Workbook
    paid_f = Payment.status.in_([PaymentStatus.PAID, PaymentStatus.PARTIAL])
    payments = (await db.execute(select(Payment).where(
        func.extract("month", Payment.created_at) == month,
        func.extract("year", Payment.created_at) == year, paid_f
    ).order_by(Payment.created_at))).scalars().all()
    expenses = (await db.execute(select(Expense).where(
        func.extract("month", Expense.expense_date) == month,
        func.extract("year", Expense.expense_date) == year
    ).order_by(Expense.expense_date))).scalars().all()

    wb = Workbook()
    # Revenue sheet
    ws1 = wb.active; ws1.title = "Revenue"
    ws1.append(["Date", "Amount (EGP)", "Method", "Receipt #"])
    for p in payments:
        ws1.append([str(p.created_at.date()), float(p.amount), p.method.value, p.receipt_number or ""])
    total_rev = sum(float(p.amount) for p in payments)
    ws1.append([]); ws1.append(["TOTAL", total_rev])

    # Expenses sheet
    ws2 = wb.create_sheet("Expenses")
    ws2.append(["Date", "Category", "Amount (EGP)", "Description"])
    for e in expenses:
        ws2.append([str(e.expense_date), e.category.value, float(e.amount), e.description or ""])
    total_exp = sum(float(e.amount) for e in expenses)
    ws2.append([]); ws2.append(["TOTAL", "", total_exp])

    # Summary sheet
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Metric", "Amount (EGP)"])
    ws3.append(["Total Revenue", total_rev])
    ws3.append(["Total Expenses", total_exp])
    ws3.append(["Net Profit", total_rev - total_exp])

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=financial_{year}_{month:02d}.xlsx"})


@router.get("/attendance/export")
async def export_attendance_excel(
    date_from: date = Query(...), date_to: date = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(REPORT_ROLES))):
    """Export attendance records to Excel."""
    from openpyxl import Workbook
    from app.models.attendance import Attendance
    records = (await db.execute(select(Attendance).where(
        func.date(Attendance.check_in) >= date_from,
        func.date(Attendance.check_in) <= date_to
    ).order_by(Attendance.check_in.desc()))).scalars().all()

    wb = Workbook(); ws = wb.active; ws.title = "Attendance"
    ws.append(["Member ID", "Check-in Time", "Method"])
    for r in records:
        ws.append([str(r.member_id), str(r.check_in), r.method.value])
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance.xlsx"})
